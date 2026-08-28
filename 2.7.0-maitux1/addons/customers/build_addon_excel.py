# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
from __future__ import unicode_literals

"""扫描 /opt/addons 下所有 addon 包，提取摘要并生成 Excel。
仅用于文档整理：从 setup.py / README / custom-addon.cfg 自动提取，不改动任何 addon。"""
import glob
import io
import os
import re
import sys

# --- openpyxl 依赖注入（从 buildout eggs 加载）---
EGGS = "/home/senaite/senaitelims/eggs/cp27mu"
if os.path.isdir(EGGS):
    for egg in sorted(glob.glob(os.path.join(EGGS, "*.egg"))):
        if egg not in sys.path:
            sys.path.insert(0, egg)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = "/opt/addons"
CFG = os.path.join(ROOT, "customers", "custom-addon.cfg")


def read_cfg():
    """解析 custom-addon.cfg，返回 {egg: {"develop":,"eggs":,"zcml":,"overrides":,"profiles":}}"""
    sec = None
    lines = {}
    if os.path.exists(CFG):
        raw = io.open(CFG, "r", encoding="utf-8").read().splitlines()
        for ln in raw:
            ln = ln.strip()
            if ln.startswith("[") and ln.endswith("]"):
                sec = ln[1:-1].strip()
                lines.setdefault(sec, [])
                continue
            if sec and ln:
                lines[sec].append(ln)
    # zcml 行里以 '-overrides' 结尾的即为 overrides 条目
    reg = {}
    for egg in lines.get("eggs", []):
        x = egg.strip()
        if not x:
            continue
        reg[x] = {
            "develop": x in [i.strip() for i in lines.get("develop", [])],
            "zcml": x in [i.strip().split("-overrides")[0] for i in lines.get("zcml", [])],
            "overrides": (x + "-overrides") in [i.strip() for i in lines.get("zcml", [])],
            "profiles": ("%s:default" % x) in [i.strip() for i in lines.get("profiles", [])],
        }
    return reg


def parse_setup(path):
    txt = io.open(path, "r", encoding="utf-8", errors="ignore").read()
    def grab(key, default=""):
        m = re.search(r'^\s*%s\s*=\s*["\'](.*?)["\']' % key, txt, re.M | re.S)
        return m.group(1).strip() if m else default
    return grab("name"), grab("version"), grab("description")


def parse_readme(path):
    txt = io.open(path, "r", encoding="utf-8", errors="ignore").read()
    out = []
    for ln in txt.splitlines():
        s = ln.strip()
        if not s or s.startswith(("#", "=", "-", "*", "<", "/")):
            continue
        s = re.sub(r"```.*", "", s)
        if 1 < len(s) < 120 and out.count(s) == 0:
            out.append(s)
            if len(out) >= 6:
                break
    return " ".join(s for s in out) if out else ""


def main():
    reg = read_cfg()
    wb = Workbook()
    ws = wb.active
    ws.title = "AddonList"
    headers = ["序号", "包名(Dist)", "Egg名", "导入包名", "版本", "目录归属",
               "功能摘要", "已注册(buildout)", "overrides", "开发规则文档"]
    ws.append(headers)
    ws.freeze_panes = "A2"
    hfill = PatternFill("solid", fgColor="4472C4")
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = []
    for setup in glob.glob(os.path.join(ROOT, "*", "*", "setup.py")) + \
                glob.glob(os.path.join(ROOT, "*", "setup.py")):
        pkg_dir = os.path.dirname(setup)
        # 目录归属：common 还是 customers
        rel = os.path.relpath(pkg_dir, ROOT)
        parts = rel.replace("\\", "/").split("/")
        belong = parts[0] if parts else ""
        name = version = desc = ""
        import_name = ""
        for cand in (os.path.join(pkg_dir, "README.md"),
                     os.path.join(pkg_dir, "README.rst")):
            if os.path.exists(cand):
                summary = parse_readme(cand)
                break
        else:
            summary = ""
        name, version, desc = parse_setup(setup)
        # 推断导入包名：与 egg 同名，去版本
        import_name = name or os.path.basename(pkg_dir)
        egg = name or ""
        r = reg.get(egg, {})
        registered = "是" if r.get("zcml") else "否"
        overrides = "是" if r.get("overrides") else "否"
        if not os.path.exists(os.path.join(pkg_dir, "overrides.zcml")) and not find_overrides(pkg_dir):
            overrides = "否(无文件)" if not r.get("overrides") else overrides
        rows.append([pkg_dir, egg, import_name, version, belong, summary, registered, overrides])

    rows.sort(key=lambda x: (x[4], x[1]))  # 按归属再按 egg 名排序
    for i, r in enumerate(rows, 1):
        ws.append([i, r[1], r[2], r[3], r[4], r[5], r[6], r[7]])
    # 对齐/宽度
    widths = [6, 30, 22, 20, 10, 12, 80, 14, 12]
    for k, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(k)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (6, 7)))

    out = os.path.join(ROOT, "customers", "Addon摘要.xlsx")
    wb.save(out)
    print("saved:", out)


def find_overrides(pkg_dir):
    return any(f.endswith("overrides.zcml") for _, _, fs in os.walk(pkg_dir) for f in fs)


if __name__ == "__main__":
    main()